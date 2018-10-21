from openpyxl import load_workbook
import Flight_data
import time


# Flight1 = Flight_data.Flight()


def read_data(fname):
    wb = load_workbook(filename=fname)
    ws = wb.get_sheet_by_name('Inputs')
    flights = []
    for row in ws.iter_rows(min_row=3, max_col=6, max_row=7):
        f_data = []
        for cell in row:
            f_data.append(cell.value)
            print(cell.value)
        flights.append(Flight_data.Flight(f_data[0], f_data[1], f_data[2], f_data[3], f_data[4], f_data[5]))

    bays = []
    for row in ws.iter_rows(min_row=12, max_row=16):
        if row[0].value == None:
            comp_list = []
            for cell in row[2:]:
                print(cell.value)
                if cell.value != None:
                    comp_list.append(cell.value)
            bays.append(Flight_data.Bay(row[1].value, False, comp_list))

        if row[1].value == None:
            comp_list = []
            for cell in row[2:]:
                print(cell.value)
                if cell.value != None:
                    comp_list.append(cell.value)
            bays.append(Flight_data.Bay(row[0].value, True, comp_list))

    occupied_bays = []
    for row in ws.iter_rows(min_row=24, max_col=1):
        for cell in row:
            if not cell.value == None: occupied_bays.append(cell.value)

    un_occupied_bays = []
    for row in ws.iter_rows(min_row=24, max_col=2, min_col=2):
        for cell in row:
            if not cell.value == None: un_occupied_bays.append(cell.value)
    print("fff")
    return flights, bays, occupied_bays, un_occupied_bays


flights, bays, occupied_bays, un_occupied_bays = read_data('data/Research_Wimali.xlsx')

######################################################################
aircraft_types = set([])
for flight in flights:
    aircraft_types.add(flight.aircraft_type)
######################################################################

# Bay compatibility of each aircraft
bay_compatability = {}
for aircraft in aircraft_types:
    bay_compatability[aircraft] = []
    for bay in bays:
        if aircraft in bay.compat:
            bay_compatability[aircraft].append(bay.bay_id)

S_bridged = {}  # A dictionary containing bridged bays for each aircraft

for aircraft in aircraft_types:
    S_bridged[aircraft] = []
    for bay in bays:
        if bay.bridged and aircraft in bay.compat:
            # add to dictionary
            S_bridged[aircraft].append(bay.bay_id)

S_unbridged = {}  # A dictionary containing unbridged bays for each aircraft

for aircraft in aircraft_types:
    S_unbridged[aircraft] = []
    for bay in bays:
        if not bay.bridged and aircraft in bay.compat:
            # add to dictionary
            S_unbridged[aircraft].append(bay.bay_id)


# generate set3 = Intersection of S_bridged & bay_compatibility

set2 = {}
for ac in aircraft_types:
    set2[ac] = list(set(S_bridged[ac]).intersection(set(bay_compatability[ac])))

def getSet2(aircraft_id):
    if not list(set(S_bridged[ac]).intersection(set(bay_compatability[ac]))): # if intersection is empty
        return list(set(S_bridged[ac]).intersection(set(bay_compatability[ac])))
    else:
        return list(set(S_unbridged[ac]).intersection(set(bay_compatability[ac])))

set2 = getSet2('A320')
print("dvck")

def getCompatibleBays(flight_name):
    bay_compat = []
    corr_aircraft = None
    for flight in flights:
        if flight.flight_name==flight_name:
            corr_aircraft = flight.aircraft_type

    for bay in bays:
        if corr_aircraft in bay.compat:
            bay_compat.append(bay.bay_id)
    return bay_compat

from constraint import *

problem = Problem()

flight_names = []
for flight in flights:
    flight_names.append(flight.flight_name)

bays_list = []
for bay in bays:
    bays_list.append(bay.bay_id)

problem.addVariables(flight_names, bays_list)

for flight in flights:
    local_compat = getCompatibleBays(flight.flight_name)
    problem.addConstraint(lambda fl: fl in local_compat.copy(), [flight.flight_name])


time_constraints = {'MI428':(1,3) , 'UL867':(4,7), 'QR664':(8,9), 'TK730':(15,16), 'UL303':(16,17)}
# problem.addConstraint(lambda fl: )
#
# def timeConstraint(fl,)

solutions = problem.getSolutions()

print (solutions)